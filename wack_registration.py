import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

camper = pd.read_csv("WACK_2025_Camper_Registration (1).csv")
sponsor = pd.read_csv("WACK_2025_Sponsor_Registration (1).csv")

camper["Name"] = camper["Child - Person's Name - First Name"] + " " + camper["Child - Person's Name - Last Name"]
camper["Gender"] = camper["Child - Gender"]
camper["Grade"] = camper["Child - Last Grade Completed"]
camper["Medical Info"] = camper["Child - Medical Information/Allergies"]
camper["Church"] = camper.apply(lambda x: x["Child - Church Name"] if x["Child - Church"] == "Other" else x["Child - Church"], axis=1)
track_map = {
    "PHOTOGRAPHY (must bring a digital camera)": "PHOTOGRAPHY",
    "STOMP (must bring drum sticks)": "STOMP",
    "Selected for Speaking Drama Role": "DRAMA"
}
camper["Track"] = camper["Child - Track Assignment"].replace(track_map)
camper["Size"] = camper["Child - T-Shirt Size"]
camper = camper[["Name", "Gender", "Grade", "Medical Info", "Church", "Track", "Size", "Email"]]

sponsor["Name"] = sponsor["Sponsor - Person's Name - First Name"] + " " + sponsor["Sponsor - Person's Name - Last Name"]
sponsor["Gender"] = sponsor["Sponsor - Gender"]
sponsor["Church"] = sponsor.apply(lambda x: x["Sponsor - Church Name"] if x["Sponsor - Church"] == "Other" else x["Sponsor - Church"], axis=1)
sponsor["Type"] = sponsor["Sponsor - "]
sponsor["Size"] = sponsor["Sponsor - T-Shirt Size"]
sponsor["Email"] = sponsor["Email"]
sponsor["Job Preference"] = sponsor["Sponsor - Preferred Job Assignment"]
sponsor["Job Details"] = sponsor["Sponsor - Preferred Job Assignment Details"]
sponsor = sponsor[["Name", "Gender", "Type", "Job Details", "Church", "Job Preference", "Size", "Email"]]

header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
header_font = Font(bold=True, size=12)

wb = openpyxl.Workbook()
wb.remove(wb.active)

for church in camper["Church"].dropna().unique():
    if not church:
        continue
    ws = wb.create_sheet(church[:31])
    
    for col, header in enumerate(camper.columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    church_campers = camper[camper["Church"] == church]
    for idx, row in enumerate(church_campers.values, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=idx, column=col, value=val)
    
    church_sponsors = sponsor[sponsor["Church"] == church]
    if len(church_sponsors) > 0:
        spacer_row = len(church_campers) + 2
        ws.cell(row=spacer_row, column=1, value="SPONSORS")
        ws.cell(row=spacer_row, column=1).font = header_font
        
        sponsor_start = spacer_row + 1
        for col, header in enumerate(sponsor.columns, 1):
            cell = ws.cell(row=sponsor_start, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for idx, row in enumerate(church_sponsors.values, sponsor_start + 1):
            for col, val in enumerate(row, 1):
                ws.cell(row=idx, column=col, value=val)
        
        size_counts = church_campers["Size"].value_counts().sort_values(ascending=False)
        size_row = sponsor_start + len(church_sponsors) + 2
        size_text = ", ".join([f"{size}: {count}" for size, count in size_counts.items()])
        ws.cell(row=size_row, column=1, value=f"T-Shirt Sizes: {size_text}")

wb.save("WACK_2025_Camper_Registration_By_Church.xlsx")

wb = openpyxl.Workbook()
wb.remove(wb.active)

for track in camper["Track"].dropna().unique():
    if not track:
        continue
    ws = wb.create_sheet(track[:31])
    
    track_campers = camper[camper["Track"] == track]
    
    for col, header in enumerate(camper.columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for idx, row in enumerate(track_campers.values, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=idx, column=col, value=val)

wb.save("WACK_2025_Camper_Registration_By_Track.xlsx")

print("Created: WACK_2025_Camper_Registration_By_Church.xlsx")
print("Created: WACK_2025_Camper_Registration_By_Track.xlsx")